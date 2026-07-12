from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ROW = 4
DATA_START_ROW = 5


def find_default_workbook() -> Path:
    downloads = Path.home() / "Downloads"
    candidates = sorted(downloads.glob("Manufacturing Daily List*.xlsx"))
    preferred = [path for path in candidates if "_" in path.name]
    if preferred:
        return preferred[0]
    if candidates:
        return candidates[0]
    raise FileNotFoundError("Could not find Manufacturing Daily List*.xlsx in Downloads.")


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return str(value).strip() or None


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y", "완료", "o", "ok"}


def parse_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", text):
        return text.replace(".", "-").replace("/", "-")
    if re.fullmatch(r"\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}", text):
        return text.replace(".", "-").replace("/", "-")
    return None


def merged_values(ws) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, col)] = value
    return values


def cell_value(ws, merged: dict[tuple[int, int], Any], row: int, col: int) -> Any:
    return merged.get((row, col), ws.cell(row, col).value)


def is_work_sheet(ws) -> bool:
    headers = [clean_text(ws.cell(HEADER_ROW, col).value) for col in range(1, 13)]
    return headers[:2] == ["No.", "업무 명"] and ws.title.upper().startswith("BAY")


def extract_records(workbook_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)
    records: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        if not is_work_sheet(ws):
            continue

        merged = merged_values(ws)
        bay = clean_text(ws["B1"].value) or ws.title

        for row in range(DATA_START_ROW, ws.max_row + 1):
            record = {
                "bay": bay,
                "source_row": row,
                "work_no": parse_int(cell_value(ws, merged, row, 1)),
                "work_name": clean_text(cell_value(ws, merged, row, 2)),
                "work_detail": clean_text(cell_value(ws, merged, row, 3)),
                "vendor": clean_text(cell_value(ws, merged, row, 4)),
                "part_no": clean_text(cell_value(ws, merged, row, 5)),
                "item_name": clean_text(cell_value(ws, merged, row, 6)),
                "bolt": clean_text(cell_value(ws, merged, row, 7)),
                "has_issue": parse_bool(cell_value(ws, merged, row, 8)),
                "worker": clean_text(cell_value(ws, merged, row, 9)),
                "work_date": parse_date(cell_value(ws, merged, row, 10)),
                "is_completed": parse_bool(cell_value(ws, merged, row, 11)),
                "issue_note": clean_text(cell_value(ws, merged, row, 12)),
            }

            meaningful_keys = [
                "work_no",
                "work_name",
                "work_detail",
                "vendor",
                "part_no",
                "item_name",
                "bolt",
                "worker",
                "work_date",
                "issue_note",
            ]
            if any(record[key] is not None for key in meaningful_keys) or record["has_issue"] or record["is_completed"]:
                records.append(record)

    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract work_items rows from Manufacturing Daily List workbook.")
    parser.add_argument("--input", type=Path, default=None)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    workbook_path = args.input or find_default_workbook()
    records = extract_records(workbook_path)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(
            {
                "source": str(workbook_path),
                "record_count": len(records),
                "records": records,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    per_bay: dict[str, int] = {}
    for record in records:
        per_bay[record["bay"]] = per_bay.get(record["bay"], 0) + 1

    print(json.dumps({"source": str(workbook_path), "record_count": len(records), "per_bay": per_bay}, ensure_ascii=False))


if __name__ == "__main__":
    main()
